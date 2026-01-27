import polars as pl
from openpyxl import load_workbook


def load_csv_lazy(path, log):
    log.info("Scanning CSV (lazy): %s", path)
    return pl.scan_csv(path)


def _make_unique_headers(header):
    seen = {}
    out = []
    for i, h in enumerate(header):
        h = "" if h is None else str(h).strip()
        if not h:
            h = "col_%d" % (i + 1)

        if h not in seen:
            seen[h] = 0
            out.append(h)
        else:
            seen[h] += 1
            out.append("%s__%d" % (h, seen[h]))
    return out


def _sheet_to_frame(ws, log, path, sheet_name, max_rows=None):
    rows_iter = ws.iter_rows(values_only=True)

    # Find first non-empty row as header
    header = None
    for r in rows_iter:
        if r is None:
            continue
        if any(x is not None and str(x).strip() != "" for x in r):
            header = list(r)
            break

    if header is None:
        log.warning("Empty sheet: %s (sheet=%s)", path.name, sheet_name)
        return pl.DataFrame()

    header = _make_unique_headers(header)
    n = len(header)

    data = []
    count = 0
    for r in rows_iter:
        if r is None:
            continue
        r = list(r)
        if len(r) < n:
            r = r + [None] * (n - len(r))
        elif len(r) > n:
            r = r[:n]
        data.append(r)
        count += 1
        if max_rows is not None and count >= max_rows:
            log.warning(
                "Excel sheet truncated at %d rows: file=%s sheet=%s",
                max_rows,
                path.name,
                sheet_name,
            )
            break

    # If there are no data rows, still return empty DF with columns
    if not data:
        return pl.DataFrame(schema=header)

    # Column-wise data
    cols = list(zip(*data))
    columns = {}
    for i in range(n):
        values = list(cols[i])

        # strict=False allows mixed types within a column
        try:
            columns[header[i]] = pl.Series(header[i], values, strict=False)
        except Exception as e:
            # worst case: coerce everything to string
            log.warning(
                "Column conversion fallback to string: file=%s sheet=%s col=%s err=%s",
                path.name, sheet_name, header[i], str(e)
            )
            columns[header[i]] = pl.Series(header[i], [None if v is None else str(v) for v in values], strict=False)

    return pl.DataFrame(columns)


def excel_to_frames(path, log, max_rows=None):
    log.info("Loading Excel: %s", path)
    wb = load_workbook(filename=path, read_only=True, data_only=True)

    out = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            out[sheet_name] = _sheet_to_frame(ws, log, path, sheet_name, max_rows=max_rows)
        except Exception as e:
            log.exception("Failed to parse sheet: file=%s sheet=%s err=%s", path.name, sheet_name, str(e))
            out[sheet_name] = pl.DataFrame()

    return out


def load_datasets(path, log, config=None):
    path = path.resolve()
    suffix = path.suffix.lower()
    cfg = config or {}

    if suffix == ".csv":
        lf = load_csv_lazy(path, log)
        yield {"path": path, "name": path.stem, "sheet": None, "df": None, "lf": lf}
        return

    if suffix in (".xlsx", ".xls"):
        max_rows = cfg.get("excel_max_rows")
        sheets = excel_to_frames(path, log, max_rows=max_rows)
        for sheet_name, df in sheets.items():
            truncated = max_rows is not None and df.height >= int(max_rows)
            yield {
                "path": path,
                "name": path.stem,
                "sheet": str(sheet_name),
                "df": df,
                "lf": None,
                "truncated": truncated,
            }
        return

    log.warning("Unsupported file type: %s", path)
